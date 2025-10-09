import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# === CONFIGURATION ===
SOURCE_FILE = "Draft IND AS  Financials_ FY 24-25.xlsx"
SOURCE_SHEET = "TB 25"
EXTRACTED_TAB = "Extracted_Data"
VALIDATION_TAB = "Validation_Report"
OUTPUT_FILE = "Processed_Balance_Sheet.xlsx"

TARGET_COLUMNS = [
    "G/L Acct/BP Code", "Name", "Local Currency - Indian Rupee - OB",
    "Local Currency - Indian Rupee - Debit", "Local Currency - Indian Rupee - Credit",
    "Local Currency - Indian Rupee - Balance", "Jeev TB", "Elimination entry",
    "Closing Balance", "Audit entries", "Net Closing Balance",
    "Function", "Grouping", "Sub-Grouping"
]

# === STEP 1: Detect Header Row ===
header_row = None
for i in range(10):
    preview = pd.read_excel(SOURCE_FILE, sheet_name=SOURCE_SHEET, skiprows=i, nrows=1, header=None)
    if not preview.empty:
        preview_values = [str(v).strip().lower() for v in preview.iloc[0].values if pd.notna(v)]
    else:
        preview_values = []
    matches = sum(any(tc.lower() in val for val in preview_values) for tc in TARGET_COLUMNS)
    if matches >= 5:
        header_row = i
        break

if header_row is None:
    print("❌ Could not detect header row. Please check the sheet format.")
    exit(1)

# === STEP 2: Load Data from Detected Header Row ===
df_raw = pd.read_excel(SOURCE_FILE, sheet_name=SOURCE_SHEET, skiprows=header_row, dtype=str, keep_default_na=False)
original_columns = df_raw.columns.tolist()

# === STEP 3: Normalize and Match Columns ===
normalized_map = {}
seen = {}
normalized_lookup = {}

for col in original_columns:
    col_str = str(col).strip().lower().replace("\n", " ").replace("  ", " ")
    if col_str in seen:
        seen[col_str] += 1
        new_name = f"{col_str}_{seen[col_str]}"
    else:
        seen[col_str] = 0
        new_name = col_str
    normalized_map[col] = new_name
    normalized_lookup[new_name] = col

df_raw.columns = list(normalized_map.values())

# === STEP 4: Match Columns to Targets ===
matched_columns = []
final_column_names = []
validation_rows = []

for target in TARGET_COLUMNS:
    target_norm = target.strip().lower().replace("\n", " ").replace("  ", " ")
    matches = [col for col in df_raw.columns if col.startswith(target_norm)]
    if matches:
        for match in matches:
            original_name = normalized_lookup.get(match, match)
            matched_columns.append(match)
            final_column_names.append(original_name)
            sample_value = df_raw[match].dropna().astype(str).head(1).values[0] if not df_raw[match].dropna().empty else ""
            position = original_columns.index(original_name) + 1 if original_name in original_columns else "-"
            validation_rows.append([
                original_name, "Yes", position,
                str(df_raw[match].dtype), sample_value,
                int(df_raw[match].isna().sum()), ""
            ])
    else:
        validation_rows.append([target, "No", "-", "-", "-", "-", "Missing"])

df_extracted = df_raw[matched_columns]
df_extracted.columns = final_column_names

# === STEP 5: Safe Overwrite Check ===
if os.path.exists(OUTPUT_FILE):
    try:
        os.remove(OUTPUT_FILE)
    except PermissionError:
        print(f"❌ File '{OUTPUT_FILE}' is currently open or locked. Please close it and try again.")
        exit(1)

# === STEP 6: Write to Excel with Formatting ===
with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    df_extracted.to_excel(writer, sheet_name=EXTRACTED_TAB, index=False)
    wb = writer.book
    ws = writer.sheets[EXTRACTED_TAB]

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    ws.freeze_panes = "A2"

    for col_idx, col in enumerate(ws.iter_cols(min_row=2), start=1):
        lengths = [len(str(cell.value)) if cell.value is not None else 0 for cell in col]
        max_length = max(lengths) if lengths else 15
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, max_length + 2)

        header = ws.cell(row=1, column=col_idx).value
        is_numeric = isinstance(header, str) and any(k in header.lower() for k in ["balance", "debit", "credit", "entries", "ob"])
        for cell in col:
            if is_numeric:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "0." + "0"*30
            else:
                cell.alignment = Alignment(horizontal="left")

    gl_col_idx = next((i + 1 for i, c in enumerate(final_column_names) if c.startswith("G/L Acct/BP Code")), None)
    if gl_col_idx:
        for cell in ws.iter_cols(min_col=gl_col_idx, max_col=gl_col_idx, min_row=2):
            for c in cell:
                c.number_format = "@"

    ws.auto_filter.ref = ws.dimensions

    ws_val = wb.create_sheet(VALIDATION_TAB)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary = [
        ["Source File", SOURCE_FILE],
        ["Tab Name", SOURCE_SHEET],
        ["Timestamp", timestamp],
        ["Total Rows Extracted", len(df_extracted)],
        ["Total Columns Extracted", len(final_column_names)],
        ["Processing Status", "Success" if final_column_names else "Failed"]
    ]
    for row in summary:
        ws_val.append(row)

    ws_val.append([])
    ws_val.append(["Column Name", "Found", "Position", "Data Type", "Sample Value", "Empty Count", "Notes"])
    for row in validation_rows:
        ws_val.append(row)

print(f"✅ Extraction complete from sheet '{SOURCE_SHEET}'. Output saved to: {OUTPUT_FILE}")